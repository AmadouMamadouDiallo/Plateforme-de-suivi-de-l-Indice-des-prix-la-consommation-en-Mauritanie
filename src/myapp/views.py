from django.shortcuts import render, get_object_or_404, redirect
from .models import Produit, TypeProduit,Wilaya,Moughataa,Commune,PointDeVente,PrixProduit,PanierProduit
from django.db.models import Count, Avg
from django.core.paginator import Paginator
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.models import User
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.core.files.storage import FileSystemStorage
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font



def login_view(request):
    """ Vue pour la connexion """
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect('home')  # Rediriger vers la page d'accueil après connexion
        else:
            messages.error(request, "Nom d'utilisateur ou mot de passe incorrect.")

    return render(request, 'login.html')


def logout_view(request):
    """ Vue pour la déconnexion """
    logout(request)
    return redirect('login')  # Rediriger vers la page de connexion


def register(request):
    """ Vue pour l'inscription """
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        password_confirm = request.POST.get('password_confirm')

        # Validation simple
        if password != password_confirm:
            messages.error(request, "Les mots de passe ne correspondent pas.")
        elif User.objects.filter(username=username).exists():
            messages.error(request, "Ce nom d'utilisateur est déjà pris.")
        elif User.objects.filter(email=email).exists():
            messages.error(request, "Cet email est déjà utilisé.")
        else:
            # Créer l'utilisateur
            user = User.objects.create_user(username=username, email=email, password=password)
            login(request, user)  # Connecter automatiquement après inscription
            return redirect('home')

    return render(request, 'register.html')


@login_required
def home(request):
    return render(request, 'index.html')
from datetime import datetime, timedelta

from django.shortcuts import render
from django.db.models import Sum, F, Count
from django.db.models.functions import ExtractMonth, ExtractYear
from datetime import datetime, timedelta
from .models import PrixProduit, Produit, PointDeVente, PanierProduit

def dashboard_inpc(request):
    # 📅 Récupérer toutes les dates disponibles et les formater en "Janvier 2024"
    dates_disponibles = PrixProduit.objects.dates('date', 'month', order='DESC')
    mois_annees = {date.strftime("%B %Y"): date.strftime('%Y-%m') for date in dates_disponibles}

    # 📌 Récupérer la date de début et de fin sélectionnée par l'utilisateur (ou valeurs par défaut)
    date_debut = request.GET.get('date_debut', list(mois_annees.values())[-1])  # Mois le plus ancien
    date_fin = request.GET.get('date_fin', list(mois_annees.values())[0])  # Mois le plus récent

    # 📌 Filtrage des prix sur la période sélectionnée
    inpc_query = PrixProduit.objects.filter(date__range=[date_debut + "-01", date_fin + "-31"])

    # 📊 Calcul de l'INPC pondéré
    inpc_calcul = inpc_query.annotate(
        poids=F('produit__panier_produits__poid')
    ).aggregate(
        inpc=Sum(F('valeur') * F('poids')) / Sum(F('poids'))
    )['inpc'] or 0  # Si aucun résultat, INPC = 0

    # 🔝 Produits influents : calcul de la contribution en pourcentage
    total_contribution = inpc_query.annotate(
        poids=F('produit__panier_produits__poid'),
        contribution=F('valeur') * F('poids')
    ).aggregate(total=Sum('contribution'))['total'] or 1  # Pour éviter la division par zéro

    top_produits = inpc_query.annotate(
        poids=F('produit__panier_produits__poid'),
        contribution=F('valeur') * F('poids')
    ).values('produit__nom').annotate(
        total_contribution=Sum('contribution')
    ).order_by('-total_contribution')[:5]

    for produit in top_produits:
        produit['pourcentage'] = (produit['total_contribution'] / total_contribution) * 100

    # 🔁 Comparaison avec le mois précédent
    date_debut_previous = datetime.strptime(date_debut, "%Y-%m")
    previous_month = (date_debut_previous.replace(day=1) - timedelta(days=1)).strftime('%Y-%m')

    inpc_previous = PrixProduit.objects.filter(
        date__year=previous_month[:4], date__month=previous_month[5:]
    ).annotate(
        poids=F('produit__panier_produits__poid')
    ).aggregate(
        inpc=Sum(F('valeur') * F('poids')) / Sum(F('poids'))
    )['inpc'] or 0

    variation_inpc = ((inpc_calcul - inpc_previous) / inpc_previous * 100) if inpc_previous != 0 else 0

    # 📊 Statistiques globales
    stats = {
        'produits': Produit.objects.count(),
        'points_de_vente': PointDeVente.objects.count(),
    }

    # 🔼 Contexte à envoyer au template
    context = {
        'mois_annees': mois_annees,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'inpc_value': inpc_calcul,
        'top_produits': top_produits,
        'variation_inpc': variation_inpc,
        'stats': stats,
    }

    return render(request, 'dashboards/dashboard_inpc.html', context)





@login_required
def dashboard(request):
    # Récupérer les statistiques générales
    stats = {
        "Wilayas": Wilaya.objects.count(),
        "Moughataas": Moughataa.objects.count(),
        "Communes": Commune.objects.count(),
        "Produits": Produit.objects.count(),
        
        
    }

    # Récupérer les données pour les graphiques
    produits_par_type = TypeProduit.objects.annotate(count=Count('produits'))
    prix_moyen_par_type = TypeProduit.objects.annotate(prix_moyen=Avg('produits__prix'))
    prix_historique = PrixProduit.objects.values('produit__nom', 'valeur', 'date')

    # Passer les données au template
    return render(request, 'dashboard.html', {
        'stats': stats,
        'produits_par_type': produits_par_type,
        'prix_moyen_par_type': prix_moyen_par_type,
        'prix_historique': list(prix_historique),
    })



def importer_wilayas(request):
    """ Importer les wilayas depuis un fichier Excel en utilisant OpenPyxl """
    if request.method == "POST" and request.FILES.get("fichier_excel"):
        fichier_excel = request.FILES["fichier_excel"]

        try:
            # Charger le fichier Excel
            wb = openpyxl.load_workbook(fichier_excel)
            sheet = wb.active  # Utiliser la première feuille

            # Vérifier l'en-tête
            header = [cell.value for cell in sheet[1]]  # Lire la première ligne (les titres)
            if "code" not in header or "nom" not in header:
                messages.error(request, "Le fichier Excel doit contenir les colonnes 'code' et 'nom'.")
                return redirect("liste_wilayas")

            # Trouver les index des colonnes
            code_index = header.index("code")
            nom_index = header.index("nom")

            # Lire les données ligne par ligne
            for row in sheet.iter_rows(min_row=2, values_only=True):
                code = str(row[code_index]).strip()
                nom = str(row[nom_index]).strip()

                if code and nom:
                    Wilaya.objects.get_or_create(code=code, nom=nom)

            messages.success(request, "Wilayas importées avec succès !")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")

    return redirect("liste_wilayas")

def exporter_wilayas(request):
    """ Exporte la liste des Wilayas sous format Excel """
    # Créer un nouveau fichier Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Wilayas"

    # Ajouter les en-têtes
    ws.append(["ID", "Code", "Nom"])

    # Remplir avec les données des wilayas
    for wilaya in Wilaya.objects.all():
        ws.append([wilaya.id, wilaya.code, wilaya.nom])

    # Préparer la réponse HTTP pour le téléchargement
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="wilayas.xlsx"'

    # Sauvegarder le fichier Excel dans la réponse
    wb.save(response)

    return response

def importer_moughataas(request):
    if request.method == "POST":
        fichier_excel = request.FILES["fichier_excel"]
        if fichier_excel.name.endswith(".xlsx"):
            wb = openpyxl.load_workbook(fichier_excel)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Ignorer l'en-tête
                code, nom, wilaya_code = row
                wilaya = Wilaya.objects.filter(code=wilaya_code).first()
                if wilaya:
                    Moughataa.objects.create(code=code, nom=nom, wilaya=wilaya)
            messages.success(request, "Importation réussie des Moughataas.")
        else:
            messages.error(request, "Veuillez importer un fichier .xlsx valide.")
    return redirect("liste_moughataas")

# 🔹 EXPORTER MOUGHATAAS
def exporter_moughataas(request):
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="moughataas.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Moughataas"
    
    # En-têtes
    ws.append(["Code", "Nom", "Wilaya"])
    
    # Données
    for moughataa in Moughataa.objects.select_related("wilaya").all():
        ws.append([moughataa.code, moughataa.nom, moughataa.wilaya.code])
    
    wb.save(response)
    return response

def importer_communes(request):
    """ Importation des communes depuis un fichier Excel (avec code Moughataa) """
    if request.method == "POST" and request.FILES.get("fichier_excel"):
        fichier_excel = request.FILES["fichier_excel"]
        workbook = openpyxl.load_workbook(fichier_excel)
        feuille = workbook.active

        try:
            for ligne in feuille.iter_rows(min_row=2, values_only=True):  # Ignorer la première ligne (entêtes)
                code, nom, code_moughataa = ligne  # Utilisation du code moughataa

                # Vérifier si la moughataa existe avec le code donné
                moughataa = Moughataa.objects.filter(code=code_moughataa).first()
                if moughataa:
                    # Créer la commune si elle n'existe pas déjà
                    Commune.objects.get_or_create(code=code, nom=nom, moughataa=moughataa)
                else:
                    messages.warning(request, f"MOUGHATAA NON TROUVÉE pour le code {code_moughataa}. Ignorée.")

            messages.success(request, "Importation réussie des communes !")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")

    return redirect("liste_communes")


def exporter_communes(request):
    """ Exportation des communes vers un fichier Excel (code moughataa inclus) """
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=communes.xlsx"

    workbook = openpyxl.Workbook()
    feuille = workbook.active
    feuille.title = "Communes"

    # Ajouter les en-têtes
    feuille.append(["Code", "Nom", "Code Moughataa"])

    # Ajouter les données
    for commune in Commune.objects.all():
        feuille.append([commune.code, commune.nom, commune.moughataa.code])

    workbook.save(response)
    return response

def importer_types_produits(request):
    """ Importer les Types de Produits depuis un fichier Excel """
    if request.method == "POST" and request.FILES.get("fichier_excel"):
        fichier_excel = request.FILES["fichier_excel"]
        wb = openpyxl.load_workbook(fichier_excel)
        feuille = wb.active

        # Lire les données à partir de la deuxième ligne (ignorer les en-têtes)
        for ligne in feuille.iter_rows(min_row=2, values_only=True):
            code, nom = ligne  # Extraction des données

            if code and nom:
                # Vérifier si le type de produit existe déjà
                if not TypeProduit.objects.filter(code=code).exists():
                    TypeProduit.objects.create(code=code, nom=nom)

        messages.success(request, "Types de Produits importés avec succès.")
        return redirect("liste_types_produits")

    return redirect("liste_types_produits")


def exporter_types_produits(request):
    """ Exporter les Types de Produits en fichier Excel """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Types de Produits"

    # Ajout des en-têtes
    ws.append(["Code", "Nom"])

    # Ajout des données
    for type_produit in TypeProduit.objects.all():
        ws.append([type_produit.code, type_produit.nom])

    # Création de la réponse HTTP
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="types_produits.xlsx"'
    wb.save(response)
    return response



# Exporter Produits en Excel
def exporter_produits(request):
    """ Exporte la liste des produits au format Excel """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produits"

    # Ajouter les en-têtes
    ws.append(["Nom", "Prix", "Description", "Unité de mesure", "Type de Produit"])

    # Ajouter les données
    for produit in Produit.objects.all():
        ws.append([
            produit.nom,
            float(produit.prix),
            produit.description if produit.description else "",
            produit.unite_mesure,
            produit.type_produit.nom
        ])

    # Réponse HTTP avec le fichier Excel
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="produits.xlsx"'
    wb.save(response)
    return response

# Importer Produits depuis Excel
def importer_produits(request):
    """ Importe une liste de produits depuis un fichier Excel """
    if request.method == "POST":
        fichier_excel = request.FILES.get("fichier_excel")
        if fichier_excel:
            try:
                wb = openpyxl.load_workbook(fichier_excel)
                ws = wb.active

                for row in ws.iter_rows(min_row=2, values_only=True):  # Ignorer la première ligne (en-têtes)
                    nom, prix, description, unite_mesure, type_produit_nom = row
                    
                    # Vérifier si le type de produit existe
                    type_produit, created = TypeProduit.objects.get_or_create(nom=type_produit_nom)

                    # Ajouter le produit
                    Produit.objects.create(
                        nom=nom,
                        prix=prix,
                        description=description,
                        unite_mesure=unite_mesure,
                        type_produit=type_produit
                    )

                messages.success(request, "Produits importés avec succès !")
            except Exception as e:
                messages.error(request, f"Erreur lors de l'importation : {e}")
        else:
            messages.error(request, "Veuillez fournir un fichier Excel valide.")

    return redirect("liste_produits")

# Exporter Produits en Excel
def exporter_produits(request):
    """ Exporte la liste des produits au format Excel """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produits"

    # Ajouter les en-têtes
    ws.append(["Nom", "Prix", "Description", "Unité de mesure", "Type de Produit"])

    # Ajouter les données
    for produit in Produit.objects.all():
        ws.append([
            produit.nom,
            float(produit.prix),
            produit.description if produit.description else "",
            produit.unite_mesure,
            produit.type_produit.nom
        ])

    # Réponse HTTP avec le fichier Excel
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="produits.xlsx"'
    wb.save(response)
    return response

# Importer Produits depuis Excel
def importer_produits(request):
    """ Importe une liste de produits depuis un fichier Excel """
    if request.method == "POST":
        fichier_excel = request.FILES.get("fichier_excel")
        if fichier_excel:
            try:
                wb = openpyxl.load_workbook(fichier_excel)
                ws = wb.active

                for row in ws.iter_rows(min_row=2, values_only=True):  # Ignorer la première ligne (en-têtes)
                    nom, prix, description, unite_mesure, type_produit_nom = row
                    
                    # Vérifier si le type de produit existe
                    type_produit, created = TypeProduit.objects.get_or_create(nom=type_produit_nom)

                    # Ajouter le produit
                    Produit.objects.create(
                        nom=nom,
                        prix=prix,
                        description=description,
                        unite_mesure=unite_mesure,
                        type_produit=type_produit
                    )

                messages.success(request, "Produits importés avec succès !")
            except Exception as e:
                messages.error(request, f"Erreur lors de l'importation : {e}")
        else:
            messages.error(request, "Veuillez fournir un fichier Excel valide.")

    return redirect("liste_produits")

def importer_points_de_vente(request):
    """ Importe des points de vente à partir d'un fichier Excel """
    if request.method == "POST" and request.FILES.get("fichier_excel"):
        fichier_excel = request.FILES["fichier_excel"]
        try:
            wb = openpyxl.load_workbook(fichier_excel)
            ws = wb.active  # Prend la première feuille

            lignes = list(ws.iter_rows(min_row=2, values_only=True))  # Ignorer l'en-tête

            for ligne in lignes:
                code, nom, type_pdv, gps_latitude, gps_longitude, commune_nom = ligne

                # Vérifier si la commune existe avant d'importer
                try:
                    commune = Commune.objects.get(nom=commune_nom)
                except Commune.DoesNotExist:
                    messages.warning(request, f"Commune '{commune_nom}' introuvable. Point de vente ignoré.")
                    continue

                # Créer un point de vente s'il n'existe pas déjà
                PointDeVente.objects.update_or_create(
                    code=code,
                    defaults={
                        "nom": nom,
                        "type": type_pdv,
                        "gps_latitude": gps_latitude,
                        "gps_longitude": gps_longitude,
                        "commune": commune
                    }
                )

            messages.success(request, "Points de vente importés avec succès.")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {str(e)}")

    return redirect("liste_points_de_vente")

def exporter_points_de_vente(request):
    """ Exporte la liste des points de vente en fichier Excel """
    # Créer un nouveau fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Points de Vente"

    # Ajouter un titre
    title_font = Font(bold=True)
    headers = ["Code", "Nom", "Type", "GPS Latitude", "GPS Longitude", "Commune"]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = title_font

    # Remplir le fichier avec les données
    for pdv in PointDeVente.objects.all():
        ws.append([
            pdv.code,
            pdv.nom,
            pdv.type,
            pdv.gps_latitude,
            pdv.gps_longitude,
            pdv.commune.nom  # Afficher le nom de la commune au lieu de son ID
        ])

    # Répondre avec un fichier Excel
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="points_de_vente.xlsx"'
    wb.save(response)
    return response

def importer_prix_produits(request):
    if request.method == "POST" and request.FILES.get("fichier_excel"):
        fichier_excel = request.FILES["fichier_excel"]
        wb = openpyxl.load_workbook(fichier_excel)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            produit_nom, point_de_vente_nom, valeur, date = row
            try:
                produit = Produit.objects.get(nom=produit_nom)
                point_de_vente = PointDeVente.objects.get(nom=point_de_vente_nom)

                PrixProduit.objects.create(
                    produit=produit,
                    point_de_vente=point_de_vente,
                    valeur=valeur,
                    date=date
                )
            except Produit.DoesNotExist:
                messages.error(request, f"Produit {produit_nom} non trouvé.")
            except PointDeVente.DoesNotExist:
                messages.error(request, f"Point de vente {point_de_vente_nom} non trouvé.")

        messages.success(request, "Importation réussie !")
        return redirect("liste_prix_produits")

    return redirect("liste_prix_produits")

# EXPORTATION PRIX PRODUITS
def exporter_prix_produits(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prix Produits"

    ws.append(["Produit", "Point de Vente", "Valeur", "Date"])
    for prix in PrixProduit.objects.all():
        ws.append([prix.produit.nom, prix.point_de_vente.nom, prix.valeur, prix.date])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="prix_produits.xlsx"'
    wb.save(response)
    return response

def importer_paniers(request):
    if request.method == "POST" and request.FILES.get("fichier_excel"):
        fichier_excel = request.FILES["fichier_excel"]
        wb = openpyxl.load_workbook(fichier_excel)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            produit_nom, date_debut, date_fin, poid = row
            try:
                produit = Produit.objects.get(nom=produit_nom)

                PanierProduit.objects.create(
                    produit=produit,
                    date_debut=date_debut,
                    date_fin=date_fin,
                    poid=poid
                )
            except Produit.DoesNotExist:
                messages.error(request, f"Produit {produit_nom} non trouvé.")

        messages.success(request, "Importation réussie !")
        return redirect("liste_paniers")

    return redirect("liste_paniers")

# EXPORTATION PANIER PRODUITS
def exporter_paniers(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Paniers"

    ws.append(["Produit", "Date Début", "Date Fin", "Poid"])
    for panier in PanierProduit.objects.all():
        ws.append([panier.produit.nom, panier.date_debut, panier.date_fin, panier.poid])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="paniers.xlsx"'
    wb.save(response)
    return response
#--------------------------------------------------------------------------




@login_required
def liste_produits(request):
    """ Affiche la liste paginée des produits (5 par page) """
    produits_list = Produit.objects.all()
    paginator = Paginator(produits_list, 6)  # Afficher 5 produits par page
    page_number = request.GET.get('page')
    produits = paginator.get_page(page_number)

    return render(request, "produits/liste.html", {"produits": produits})


def ajouter_produit(request):
    """ Ajoute un nouveau produit """
    types_produits = TypeProduit.objects.all()

    if request.method == "POST":
        nom = request.POST.get("nom")
        prix = request.POST.get("prix")
        description = request.POST.get("description")
        unite_mesure = request.POST.get("unite_mesure")
        type_produit_id = request.POST.get("type_produit")

        if nom and prix and type_produit_id:
            type_produit = get_object_or_404(TypeProduit, id=type_produit_id)
            Produit.objects.create(
                nom=nom,
                prix=prix,
                description=description,
                unite_mesure=unite_mesure,
                type_produit=type_produit
            )
            messages.success(request, "Produit ajouté avec succès.")
            return redirect("liste_produits")
        else:
            messages.error(request, "Tous les champs obligatoires doivent être remplis.")

    return render(request, "produits/ajouter.html", {"types_produits": types_produits})


def modifier_produit(request, pk):
    """ Modifie un produit existant """
    produit = get_object_or_404(Produit, pk=pk)
    types_produits = TypeProduit.objects.all()

    if request.method == "POST":
        nom = request.POST.get("nom")
        prix = request.POST.get("prix")
        description = request.POST.get("description")
        unite_mesure = request.POST.get("unite_mesure")
        type_produit_id = request.POST.get("type_produit")

        if nom and prix and type_produit_id:
            produit.nom = nom
            produit.prix = prix
            produit.description = description
            produit.unite_mesure = unite_mesure
            produit.type_produit = get_object_or_404(TypeProduit, id=type_produit_id)
            produit.save()
            messages.success(request, "Produit modifié avec succès.")
            return redirect("liste_produits")
        else:
            messages.error(request, "Tous les champs obligatoires doivent être remplis.")

    return render(request, "produits/modifier.html", {"produit": produit, "types_produits": types_produits})


def supprimer_produit(request, pk):
    """ Supprime un produit """
    produit = get_object_or_404(Produit, pk=pk)

    if request.method == "POST":
        produit.delete()
        messages.success(request, "Produit supprimé avec succès.")
        return redirect("liste_produits")

    return render(request, "produits/supprimer.html", {"produit": produit})


def liste_types_produits(request):
    """ Affiche la liste paginée des types de produits (5 par page) """
    types_list = TypeProduit.objects.all()
    paginator = Paginator(types_list, 5)  # Afficher 5 types de produits par page
    page_number = request.GET.get('page')
    types_produits = paginator.get_page(page_number)

    return render(request, "types_produits/liste.html", {"types_produits": types_produits})

def ajouter_type_produit(request):
    """ Ajoute un nouveau type de produit """
    if request.method == "POST":
        code = request.POST.get("code")
        nom = request.POST.get("nom")

        if code and nom:
            TypeProduit.objects.create(code=code, nom=nom)
            messages.success(request, "Type de produit ajouté avec succès.")
            return redirect("liste_types_produits")
        else:
            messages.error(request, "Tous les champs sont obligatoires.")

    return render(request, "types_produits/ajouter.html")


def modifier_type_produit(request, pk):
    """ Modifie un type de produit existant """
    type_produit = get_object_or_404(TypeProduit, pk=pk)

    if request.method == "POST":
        code = request.POST.get("code")
        nom = request.POST.get("nom")

        if code and nom:
            type_produit.code = code
            type_produit.nom = nom
            type_produit.save()
            messages.success(request, "Type de produit modifié avec succès.")
            return redirect("liste_types_produits")
        else:
            messages.error(request, "Tous les champs sont obligatoires.")

    return render(request, "types_produits/modifier.html", {"type_produit": type_produit})


def supprimer_type_produit(request, pk):
    """ Supprime un type de produit """
    type_produit = get_object_or_404(TypeProduit, pk=pk)

    if request.method == "POST":
        type_produit.delete()
        messages.success(request, "Type de produit supprimé avec succès.")
        return redirect("liste_types_produits")

    return render(request, "types_produits/supprimer.html", {"type_produit": type_produit})

#------------------------------------------------------------------------------------
def liste_wilayas(request):
    """ Affiche la liste paginée des wilayas (5 par page) """
    wilayas_list = Wilaya.objects.all()
    paginator = Paginator(wilayas_list, 5)  # Afficher 5 wilayas par page
    page_number = request.GET.get('page')
    wilayas = paginator.get_page(page_number)

    return render(request, "wilayas/liste.html", {"wilayas": wilayas})


def ajouter_wilaya(request):
    """ Ajoute une nouvelle wilaya """
    if request.method == "POST":
        code = request.POST.get("code")
        nom = request.POST.get("nom")

        if code and nom:
            Wilaya.objects.create(code=code, nom=nom)
            messages.success(request, "Wilaya ajoutée avec succès.")
            return redirect("liste_wilayas")
        else:
            messages.error(request, "Tous les champs sont obligatoires.")

    return render(request, "wilayas/ajouter.html")


def modifier_wilaya(request, pk):
    """ Modifie une wilaya existante """
    wilaya = get_object_or_404(Wilaya, pk=pk)

    if request.method == "POST":
        code = request.POST.get("code")
        nom = request.POST.get("nom")

        if code and nom:
            wilaya.code = code
            wilaya.nom = nom
            wilaya.save()
            messages.success(request, "Wilaya modifiée avec succès.")
            return redirect("liste_wilayas")
        else:
            messages.error(request, "Tous les champs sont obligatoires.")

    return render(request, "wilayas/modifier.html", {"wilaya": wilaya})


def supprimer_wilaya(request, pk):
    """ Supprime une wilaya """
    wilaya = get_object_or_404(Wilaya, pk=pk)

    if request.method == "POST":
        wilaya.delete()
        messages.success(request, "Wilaya supprimée avec succès.")
        return redirect("liste_wilayas")

    return render(request, "wilayas/supprimer.html", {"wilaya": wilaya})

#--------------------------------------------------------------------------------
@login_required
def liste_moughataas(request):
    """ Affiche la liste paginée des Moughataas """
    moughataas_list = Moughataa.objects.all()
    paginator = Paginator(moughataas_list, 5)
    page_number = request.GET.get('page')
    moughataas = paginator.get_page(page_number)
    wilayas = Wilaya.objects.all()  # Récupérer toutes les wilayas pour le modal d'ajout

    return render(request, "moughataas/liste.html", {"moughataas": moughataas, "wilayas": wilayas})


@login_required
def ajouter_moughataa(request):
    """ Ajoute une nouvelle Moughataa """
    wilayas = Wilaya.objects.all()

    if request.method == "POST":
        code = request.POST.get("code")
        nom = request.POST.get("nom")
        wilaya_id = request.POST.get("wilaya")

        if code and nom and wilaya_id:
            wilaya = get_object_or_404(Wilaya, id=wilaya_id)
            Moughataa.objects.create(code=code, nom=nom, wilaya=wilaya)
            messages.success(request, "Moughataa ajoutée avec succès.")
            return redirect("liste_moughataas")
        else:
            messages.error(request, "Tous les champs sont obligatoires.")

    return render(request, "moughataas/ajouter.html", {"wilayas": wilayas})


@login_required
def modifier_moughataa(request, pk):
    """ Modifie une Moughataa existante """
    moughataa = get_object_or_404(Moughataa, pk=pk)
    wilayas = Wilaya.objects.all()

    if request.method == "POST":
        code = request.POST.get("code")
        nom = request.POST.get("nom")
        wilaya_id = request.POST.get("wilaya")

        if code and nom and wilaya_id:
            moughataa.code = code
            moughataa.nom = nom
            moughataa.wilaya = get_object_or_404(Wilaya, id=wilaya_id)
            moughataa.save()
            messages.success(request, "Moughataa modifiée avec succès.")
            return redirect("liste_moughataas")
        else:
            messages.error(request, "Tous les champs sont obligatoires.")

    return render(request, "moughataas/modifier.html", {"moughataa": moughataa, "wilayas": wilayas})


@login_required
def supprimer_moughataa(request, pk):
    """ Supprime une Moughataa """
    moughataa = get_object_or_404(Moughataa, pk=pk)

    if request.method == "POST":
        moughataa.delete()
        messages.success(request, "Moughataa supprimée avec succès.")
        return redirect("liste_moughataas")

    return render(request, "moughataas/supprimer.html", {"moughataa": moughataa})

#------------------------------------------------------------------------------

@login_required
def liste_communes(request):
    """ Affiche la liste paginée des Communes """
    communes_list = Commune.objects.all()
    paginator = Paginator(communes_list, 5)  # 5 Communes par page
    page_number = request.GET.get('page')
    communes = paginator.get_page(page_number)
    
    moughataas = Moughataa.objects.all()  # Liste des Moughataas pour le formulaire

    return render(request, "communes/liste.html", {"communes": communes, "moughataas": moughataas})

@login_required
def ajouter_commune(request):
    """ Ajoute une nouvelle Commune """
    if request.method == "POST":
        code = request.POST.get("code")
        nom = request.POST.get("nom")
        moughataa_id = request.POST.get("moughataa")

        if code and nom and moughataa_id:
            moughataa = get_object_or_404(Moughataa, id=moughataa_id)
            Commune.objects.create(code=code, nom=nom, moughataa=moughataa)
            messages.success(request, "Commune ajoutée avec succès.")
            return redirect("liste_communes")
        else:
            messages.error(request, "Tous les champs sont obligatoires.")

    return render(request, "communes/ajouter.html", {"moughataas": Moughataa.objects.all()})

@login_required
def modifier_commune(request, pk):
    """ Modifie une Commune existante """
    commune = get_object_or_404(Commune, pk=pk)
    moughataas = Moughataa.objects.all()

    if request.method == "POST":
        code = request.POST.get("code")
        nom = request.POST.get("nom")
        moughataa_id = request.POST.get("moughataa")

        if code and nom and moughataa_id:
            commune.code = code
            commune.nom = nom
            commune.moughataa = get_object_or_404(Moughataa, id=moughataa_id)
            commune.save()
            messages.success(request, "Commune modifiée avec succès.")
            return redirect("liste_communes")
        else:
            messages.error(request, "Tous les champs sont obligatoires.")

    return render(request, "communes/modifier.html", {"commune": commune, "moughataas": moughataas})

@login_required
def supprimer_commune(request, pk):
    """ Supprime une Commune """
    commune = get_object_or_404(Commune, pk=pk)

    if request.method == "POST":
        commune.delete()
        messages.success(request, "Commune supprimée avec succès.")
        return redirect("liste_communes")

    return render(request, "communes/supprimer.html", {"commune": commune})

#-----------------------------------------------------------------------------------

@login_required
def liste_points_de_vente(request):
    """ Affiche la liste paginée des points de vente """
    points_list = PointDeVente.objects.all()
    paginator = Paginator(points_list, 5)  # 5 points de vente par page
    page_number = request.GET.get('page')
    points_de_vente = paginator.get_page(page_number)
    communes = Commune.objects.all()

    return render(request, "points_de_vente/liste.html", {"points_de_vente": points_de_vente, "communes": communes})


@login_required
def ajouter_point_de_vente(request):
    """ Ajoute un nouveau point de vente """
    communes = Commune.objects.all()

    if request.method == "POST":
        code = request.POST.get("code")
        nom = request.POST.get("nom")
        type_point = request.POST.get("type")
        gps_latitude = request.POST.get("gps_latitude")
        gps_longitude = request.POST.get("gps_longitude")
        commune_id = request.POST.get("commune")

        if code and nom and type_point and gps_latitude and gps_longitude and commune_id:
            commune = get_object_or_404(Commune, id=commune_id)
            PointDeVente.objects.create(
                code=code,
                nom=nom,
                type=type_point,
                gps_latitude=gps_latitude,
                gps_longitude=gps_longitude,
                commune=commune
            )
            messages.success(request, "Point de vente ajouté avec succès.")
            return redirect("liste_points_de_vente")
        else:
            messages.error(request, "Tous les champs sont obligatoires.")

    return render(request, "points_de_vente/ajouter.html", {"communes": communes})


@login_required
def modifier_point_de_vente(request, pk):
    """ Modifie un point de vente existant """
    point_de_vente = get_object_or_404(PointDeVente, pk=pk)
    communes = Commune.objects.all()

    if request.method == "POST":
        code = request.POST.get("code")
        nom = request.POST.get("nom")
        type_point = request.POST.get("type")
        gps_latitude = request.POST.get("gps_latitude")
        gps_longitude = request.POST.get("gps_longitude")
        commune_id = request.POST.get("commune")

        if code and nom and type_point and gps_latitude and gps_longitude and commune_id:
            point_de_vente.code = code
            point_de_vente.nom = nom
            point_de_vente.type = type_point
            point_de_vente.gps_latitude = gps_latitude
            point_de_vente.gps_longitude = gps_longitude
            point_de_vente.commune = get_object_or_404(Commune, id=commune_id)
            point_de_vente.save()
            messages.success(request, "Point de vente modifié avec succès.")
            return redirect("liste_points_de_vente")
        else:
            messages.error(request, "Tous les champs sont obligatoires.")

    return render(request, "points_de_vente/modifier.html", {"point_de_vente": point_de_vente, "communes": communes})


@login_required
def supprimer_point_de_vente(request, pk):
    """ Supprime un point de vente """
    point_de_vente = get_object_or_404(PointDeVente, pk=pk)

    if request.method == "POST":
        point_de_vente.delete()
        messages.success(request, "Point de vente supprimé avec succès.")
        return redirect("liste_points_de_vente")

    return render(request, "points_de_vente/supprimer.html", {"point_de_vente": point_de_vente})

#-------------------------------------------------------------------------------------------
@login_required
def liste_prix_produits(request):
    """ Affiche la liste paginée des prix des produits """
    prix_list = PrixProduit.objects.all()
    paginator = Paginator(prix_list, 5)  # 5 prix par page
    page_number = request.GET.get('page')
    prix_produits = paginator.get_page(page_number)
    produits = Produit.objects.all()
    points_de_vente = PointDeVente.objects.all()

    return render(request, "prix_produits/liste.html", {"prix_produits": prix_produits, "produits": produits, "points_de_vente": points_de_vente})


@login_required
def ajouter_prix_produit(request):
    """ Ajoute un nouveau prix pour un produit """
    produits = Produit.objects.all()
    points_de_vente = PointDeVente.objects.all()

    if request.method == "POST":
        produit_id = request.POST.get("produit")
        point_de_vente_id = request.POST.get("point_de_vente")
        valeur = request.POST.get("valeur")
        date = request.POST.get("date")

        if produit_id and point_de_vente_id and valeur and date:
            produit = get_object_or_404(Produit, id=produit_id)
            point_de_vente = get_object_or_404(PointDeVente, id=point_de_vente_id)
            PrixProduit.objects.create(
                produit=produit,
                point_de_vente=point_de_vente,
                valeur=valeur,
                date=date
            )
            messages.success(request, "Prix ajouté avec succès.")
            return redirect("liste_prix_produits")
        else:
            messages.error(request, "Tous les champs sont obligatoires.")

    return render(request, "prix_produits/ajouter.html", {"produits": produits, "points_de_vente": points_de_vente})


@login_required
def modifier_prix_produit(request, pk):
    """ Modifie un prix de produit existant """
    prix_produit = get_object_or_404(PrixProduit, pk=pk)
    produits = Produit.objects.all()
    points_de_vente = PointDeVente.objects.all()

    if request.method == "POST":
        produit_id = request.POST.get("produit")
        point_de_vente_id = request.POST.get("point_de_vente")
        valeur = request.POST.get("valeur")
        date = request.POST.get("date")

        if produit_id and point_de_vente_id and valeur and date:
            prix_produit.produit = get_object_or_404(Produit, id=produit_id)
            prix_produit.point_de_vente = get_object_or_404(PointDeVente, id=point_de_vente_id)
            prix_produit.valeur = valeur
            prix_produit.date = date
            prix_produit.save()
            messages.success(request, "Prix modifié avec succès.")
            return redirect("liste_prix_produits")
        else:
            messages.error(request, "Tous les champs sont obligatoires.")

    return render(request, "prix_produits/modifier.html", {"prix_produit": prix_produit, "produits": produits, "points_de_vente": points_de_vente})


@login_required
def supprimer_prix_produit(request, pk):
    """ Supprime un prix de produit """
    prix_produit = get_object_or_404(PrixProduit, pk=pk)

    if request.method == "POST":
        prix_produit.delete()
        messages.success(request, "Prix supprimé avec succès.")
        return redirect("liste_prix_produits")

    return render(request, "prix_produits/supprimer.html", {"prix_produit": prix_produit})

#------------------------------------------------------------------------------------------------
@login_required
def liste_paniers(request):
    """ Affiche la liste paginée des paniers de produits """
    paniers_list = PanierProduit.objects.all()
    paginator = Paginator(paniers_list, 5)  # Pagination (5 par page)
    page_number = request.GET.get('page')
    paniers = paginator.get_page(page_number)

    produits = Produit.objects.all()
    
    return render(request, "paniers/liste.html", {"paniers": paniers, "produits": produits})

@login_required
def ajouter_panier(request):
    """ Ajoute un panier de produits """
    if request.method == "POST":
        produit_id = request.POST.get("produit")
        date_debut = request.POST.get("date_debut")
        date_fin = request.POST.get("date_fin")
        poid = request.POST.get("poid")

        if produit_id and date_debut and poid:
            produit = get_object_or_404(Produit, id=produit_id)
            PanierProduit.objects.create(
                produit=produit,
                date_debut=date_debut,
                date_fin=date_fin if date_fin else None,
                poid=poid
            )
            messages.success(request, "Panier ajouté avec succès.")
            return redirect("liste_paniers")

    produits = Produit.objects.all()
    return render(request, "paniers/ajouter.html", {"produits": produits})

@login_required
def modifier_panier(request, pk):
    """ Modifie un panier existant """
    panier = get_object_or_404(PanierProduit, pk=pk)

    if request.method == "POST":
        produit_id = request.POST.get("produit")
        date_debut = request.POST.get("date_debut")
        date_fin = request.POST.get("date_fin")
        poid = request.POST.get("poid")

        if produit_id and date_debut and poid:
            panier.produit = get_object_or_404(Produit, id=produit_id)
            panier.date_debut = date_debut
            panier.date_fin = date_fin if date_fin else None
            panier.poid = poid
            panier.save()
            messages.success(request, "Panier modifié avec succès.")
            return redirect("liste_paniers")

    produits = Produit.objects.all()
    return render(request, "paniers/modifier.html", {"panier": panier, "produits": produits})

@login_required
def supprimer_panier(request, pk):
    """ Supprime un panier """
    panier = get_object_or_404(PanierProduit, pk=pk)

    if request.method == "POST":
        panier.delete()
        messages.success(request, "Panier supprimé avec succès.")
        return redirect("liste_paniers")

    return render(request, "paniers/supprimer.html", {"panier": panier})

#-=================================================================================

from django.shortcuts import render
from django.db.models import Count, Avg
from .models import Produit, TypeProduit, PrixProduit, PanierProduit
from django.http import JsonResponse

def dashboard_produits(request):
    type_selectionne = request.GET.get('type', 'Tous')

    produits_query = Produit.objects.all()
    if type_selectionne != 'Tous':
        produits_query = produits_query.filter(type_produit__nom=type_selectionne)

    produits_par_type = TypeProduit.objects.annotate(count=Count('produits'))

    context = {
        'types_produits': TypeProduit.objects.values_list('nom', flat=True),
        'produits_par_type': list(produits_par_type.values('nom', 'count')),
        'produits': list(produits_query.values('nom', 'prix')),
        'type_selectionne': type_selectionne
    }
    return render(request, 'dashboards/dashboard_produits.html', context)

from django.db.models.functions import ExtractYear

def dashboard_prix_produits(request):
    annee_selectionnee = request.GET.get('annee', 'Tous')
    type_selectionne = request.GET.get('type_produit', 'Tous')

    prix_query = PrixProduit.objects.annotate(annee=ExtractYear('date'))

    if annee_selectionnee != 'Tous':
        prix_query = prix_query.filter(annee=annee_selectionnee)

    if type_selectionne != 'Tous':
        prix_query = prix_query.filter(produit__type_produit__nom=type_selectionne)

    prix_moyen_par_type = TypeProduit.objects.annotate(prix_moyen=Avg('produits__prix'))

    context = {
        'annees': list(PrixProduit.objects.dates('date', 'year').values_list('date__year', flat=True)),
        'types_produits': list(TypeProduit.objects.values_list('nom', flat=True)),
        'prix_moyen_par_type': list(prix_moyen_par_type.values('nom', 'prix_moyen')),
        'prix_produits': list(prix_query.values('produit__nom', 'valeur', 'annee')),
        'annee_selectionnee': annee_selectionnee,
        'type_selectionne': type_selectionne
    }
    return render(request, 'dashboards/dashboard_prix_produits.html', context)
def dashboard_paniers(request):
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')

    paniers_query = PanierProduit.objects.all()
    if date_debut:
        paniers_query = paniers_query.filter(date_debut__gte=date_debut)
    if date_fin:
        paniers_query = paniers_query.filter(date_fin__lte=date_fin)

    paniers_par_produit = paniers_query.values('produit__nom').annotate(count=Count('id'))

    context = {
        'paniers_par_produit': list(paniers_par_produit),
        'date_debut': date_debut,
        'date_fin': date_fin
    }
    return render(request, 'dashboards/dashboard_paniers.html', context)
